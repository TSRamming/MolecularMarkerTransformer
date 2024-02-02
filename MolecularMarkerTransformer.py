import openpyxl
import json
import csv
from pathlib import Path


def build_hierarchy_from_treelist(treelist, result, indentation=0, group=None):
    filtered_treelist = (value for value in treelist if value[2] == group)
    for treelist_item in filtered_treelist:
        result.append([indentation, treelist_item[0]])
        build_hierarchy_from_treelist(treelist, result, indentation+1, treelist_item[1])


def generate_textfile_hierarchy(spec_entries, filename):
    values = []
    for spec_entry in spec_entries:
        synonyms = ", ".join(sorted(spec_entry["Synonyms"]))
        text = spec_entry["DisplayName"]
        additional_info = (spec_entry["Code"] if spec_entry["IsHGNC"] else "na") +\
            (" / " + synonyms if synonyms != "" else "")
        if additional_info != "na":
            text += " [" + additional_info + "]"
        values.append((text, spec_entry["Code"], spec_entry["InGroup"]))
    values.sort(key=(lambda x: x[0].lower()))

    result = []
    build_hierarchy_from_treelist(values, result)
    with open(filename, 'w', encoding="utf-8") as cFile:
        cFile.write("\n".join([item[0] * "    " + item[1] for item in result]))


def generate_csv_secutrial(spec_entries, filename):
    values = []
    for spec_entry in spec_entries:
        if not isinstance(spec_entry["OtherCategory"], str) or spec_entry["OtherCategory"].strip() != "Gruppe":
            symbol = spec_entry["Code"]
            description = spec_entry["DisplayName"]
            if len(spec_entry["Synonyms"]) > 0:
                description = description + " (" + ", ".join(sorted(spec_entry["Synonyms"], key=str.lower)) + ")"
            values.append((symbol, description))
    values.sort(key=(lambda x: x[0].lower()))

    with open(filename, 'w', encoding="utf-8", newline="") as cFile:
        csv_writer = csv.writer(cFile, delimiter=";")
        csv_writer.writerow(["Ebene 1", ""])
        csv_writer.writerow(["Ebene 1: Spalte 1", "Ebene 1: Spalte 2"])
        for row in values:
            csv_writer.writerow(row)


def generate_csv_studystar(spec_entries, target_filename):
    # build import file for studystar
    # use an onkostar export xml as shell for entries
    values = []
    for spec_entry in spec_entries:
        entry_for_group = get_entry_for_code(spec_entries, spec_entry["InGroup"])
        assigned_group = ""
        if entry_for_group is not None:
            assigned_group = entry_for_group["Code"]
        values.append(
            (
                spec_entry["Code"],
                spec_entry["DisplayName"],
                spec_entry["DisplayName"],
                ",".join(sorted(spec_entry["Synonyms"])),
                "",
                assigned_group
            )
        )
    with open(target_filename, 'w', encoding="utf-8", newline="") as cFile:
        csv_writer = csv.writer(cFile, delimiter=";")
        for row in values:
            print(row)
            csv_writer.writerow(row)


def get_entry_for_code(spec_entries, code):
    """
    searches the entry matching the given code
    :param spec_entries: list of entries
    :param code: code of the entry you want returned
    :return: found entry or None
    """
    for spec_entry in spec_entries:
        if code == spec_entry['Code']:
            return spec_entry
    return None


def generate_xlsx_gravity(spec_entries, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        ["eingang", "eingangDesc", "ausgang", "ausgangDesc", "bemerkung", "entwurf", "anzahl", "systeme", "markieren",
         "markierungshinweis"]
    )
    for spec_entry in spec_entries:
        rowdata = [spec_entry["Code"], "", spec_entry["Code"], "", spec_entry["DisplayName"], "0", "", "", "0", ""]
        ws.append(rowdata)
    wb.save(filename)


def get_spec_from_worksheet(worksheet):
    entries = []
    for row in worksheet.iter_rows(2, worksheet.max_row):
        synonyms = set()
        if row[3].value:
            synonyms.update({item.strip() for item in row[3].value.split("|")})
        entry = {
            "DisplayName": row[0].value,
            "Code": row[1].value,
            "IsHGNC": row[2].value == "ja",
            "Synonyms": synonyms,
            "InGroup": row[4].value,
            "Comment": row[7].value,
            "OtherCategory": row[5].value
        }
        entries.append(entry)
    return entries


def add_synonyms_from_hgnc(spec_entries, hgnc_filename):
    """
    adds synonyms from a hgnc json dataset to entries

    :param spec_entries: the list of entries the synonyms are added to
    :param hgnc_filename: filename of the json file
    """
    # get data from hgnc json
    with open(hgnc_filename, 'r', encoding="UTF8") as file:
        hgnc = json.load(file)
    gene_list = hgnc["response"]["docs"]

    for spec_entry in spec_entries:
        if spec_entry["IsHGNC"]:
            aliases = [value.get("alias_symbol", []) for value in gene_list if value["symbol"] == spec_entry["Code"]]
            for alias in aliases:
                spec_entry["Synonyms"].update(alias)


def main():
    source_folder = "sources"
    source_excel = "MolMarker_Kategorisiert.xlsx"
    hgnc_export = "hgnc_complete_set.json"

    result_folder = "generated_files"
    hierarchy_visualisation = "visualization.txt"
    studystar_import = "import_file_molMarker_Studystar.csv"
    gravity_import = "gravity.xlsx"
    secutrial_catalogue_import = "secutrial_catalogue.csv"

    # get data from Excel file
    workbook = openpyxl.load_workbook(Path(source_folder, source_excel))
    worksheet = workbook.active
    spec_entries = get_spec_from_worksheet(worksheet)
    add_synonyms_from_hgnc(spec_entries, Path(source_folder, hgnc_export))

    generate_textfile_hierarchy(spec_entries, Path(result_folder, hierarchy_visualisation))
    generate_csv_studystar(spec_entries, Path(result_folder, studystar_import))
    generate_xlsx_gravity(spec_entries, Path(result_folder, gravity_import))
    generate_csv_secutrial(spec_entries, Path(result_folder, secutrial_catalogue_import))


if __name__ == "__main__":
    main()
